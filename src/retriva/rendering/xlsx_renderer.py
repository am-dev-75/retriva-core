# Copyright (C) 2026 Andrea Marson (am.dev.75@gmail.com)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#         http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from pathlib import Path
from typing import Dict, Optional
from retriva.rendering import register_renderer
from retriva.logger import get_logger

logger = get_logger(__name__)

class XlsxRenderer:
    """Renderer for Excel (.xlsx) artifacts using openpyxl."""

    def render(
        self,
        artifact_type: str,
        parameters: Dict[str, str],
        output_path: Path,
        cancel_check: Optional[callable] = None,
    ) -> bool:
        logger.info(f"Rendering xlsx artifact: {artifact_type}")
        
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl not installed. Cannot render xlsx.")
            return False

        title = parameters.get("title", "Retriva Artifact")
        content = parameters.get("content", "No content provided.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Artifact"
        
        ws["A1"] = title
        ws["A2"] = content
        
        if cancel_check and cancel_check():
            return False
            
        wb.save(str(output_path))
        return True

# Register the renderer
register_renderer("xlsx", XlsxRenderer)
